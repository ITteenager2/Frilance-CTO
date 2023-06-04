#Блок ввода первоначальных данных и инициализация переменных.
#############С ЛЮБОВЬЮ ЕВГЕНИЙ###############
#############С ЛЮБОВЬЮ ЕВГЕНИЙ###############
#############С ЛЮБОВЬЮ ЕВГЕНИЙ###############
import random
import math
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('tkagg')
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from scipy.interpolate import make_interp_spline

def calc_l(n):# Расчет количества исправных линий
    y3 = np.random.normal(n-1, 1)
    return int(y3)


def calc_tob(): # Расчет времени обслуживания автомобился tob распределено по закону Релея
    return 1/1.5 * math.sqrt(-math.log(random.random()))


def calc_tm(): # Расчет времени ожидания автомобилей в очереди
    return -1/0.75 * math.log(random.random())


def calc_tk():
    return 1/3.3 * math.pow(-math.log(random.random()), 1/1.5)


Xmax = 0
Tq = 0 # Срденее время ожидания в очереди
Dq, sq, Asq, Ekg = 0, 0, 0, 0 # статистические характеристики Tq
Xo = 0 # Производительность станции
Do, so, Aso, Eko = 0, 0, 0, 0 # Статистические характеристики Xo
Xn = 0# Среднее число автомобилей покинувших станицю негбслу-женными
Dn, sb, Asq, Ekq, tob, tk, tm, l = 0, 0, 0, 0, 0, 0, 0, 0 #Статистические характеристики Xn
T = 0 # Текущее время работы станции
q = 0 # Число автомобилей, ожидавших в очереди
N = 0 # Число экспериментов
n = 0 # число линий
Tmax = 0 # предельное время ожидания в очереди
ap = 0 # Математическое ожидание интервала времени прибытия автомобиля
ao = 0 # Математическое ожидание времени обслуживания автомо-биля
aq = 0 # Математическое ожидание времени ожидания в очереди
al = 0 # Математическое ожидание числа исправных линий
sl = 0 # Соответствующее среднее квадратическое отклонение
l = 0 # Число исправных линий
Tmin = 0 # Минимальное время ожидания в очереди

# Ввод данных
N = int(input('Число экспериментов: '))
n = int(input('Число линий: '))
Tmax = float(input('Предельное время ожидания в очереди: '))
ap = float(input('Математическое ожидание интервала времени прибытия автомобиля: '))
ao = float(input('Математическое ожидание времени обслуживания автомобиля: '))
aq = float(input('Математическое ожидание времени ожидания в очереди: '))
al = float(input('Математическое ожидание числа исправных линий: '))
sl = float(input('Соответствующее среднее квадратическое отклонение: '))


X = np.zeros((3, 1000))# 0 - число обслуженных автомобилей, 1 - число необслуженных автомобилей, 2 - среднее время ожидания в очереди
Ln = np.zeros(n) # Массив линий содержащий время очередного освобождения каждой линии
Freq = np.zeros((11, 4)) # Массив частот
Stat = np.zeros(n) # Массив вспомогательных статистических значений
Xsum = np.zeros((3, 4))# Массив сумм

Stat[0] = 10000
Xmin = 10000




#Блок моделирования работы системы.
unprocessed_cars = np.zeros(N)  # Массив необслуженных автомобилей
counter = 0
q = 0
l = calc_l(n)
print("Число исправных линий = ", l)
for i in range(N):
    T, q, k = 0, 0, 0
    for j in range(n):
        Ln[j] = 0
    while(T <= 630):
        j = 0
        while(j < l and T < Ln[j]):
            j += 1
            #print("T = ", T, "; Ln[j] = ", Ln[j], "; j = ", j)
        if(j < l):
            tob = calc_tob()
            X[0][i] = int(X[0][i]) + int(1)
            Ln[j] = T + tob
            k += 1
            tk = calc_tk()
            T = T + tk
        else:
            tm = calc_tm()
            if(tm >= Tmax):
                X[1][i] += 1
                unprocessed_cars[i] += 1
            else:
                X[0][i] = int(X[0][i]) + int(1)
                q += 1
                X[2][i] += tm
                #print(q)
            k += 1
            tk = calc_tk()
            T = T + tk
        #print("tm = ",tm,"; Tmax = ",Tmax)
        #print("i = ",i,"; q = ",q)
    X[2][i] = int(X[2][i])/q
    if(X[0][i] > Xmax):
        Xmax = X[0][i]
    if(X[0][i] < Xmin):
        Xmin = X[0][i]

    Xsum[0][0] = int(Xsum[0][0]) + int(X[0][i])
    Xsum[0][1] += X[0][i]*X[0][i]
    Xsum[0][2] += X[0][i]**3
    Xsum[0][3] += X[0][i]**4
    Xsum[1][0] += X[1][i]
    Xsum[1][1] += X[1][i]**2
    Xsum[1][2] += X[1][i]**3
    Xsum[1][3] += X[1][i]**4
    Xsum[2][0] += X[2][i]
    Xsum[2][1] += X[2][i]**2
    Xsum[2][2] += X[2][i]**3
    Xsum[2][3] += X[2][i]**4
    counter += 1

#Блок обработки результатов моделирования.

Xo = Xsum[0][0]/1000
Do = Xsum[0][1]/1000 - Xo * Xo
so = math.sqrt(Do)
Aso = (Xsum[0][2]/1000 - 3*Xo*Xsum[0][1]/1000 + 2*Xo**3)/(so**3)
#Eko = (Xsum[0][3]/1000 - 4*Xo*Xsum[0][2]/1000 + 6*(Xo**2)*Xsum[0][1]/1000 - 3*Xo)/(so**4)-3
for i in range(N):
    Eko += math.pow(X[0][i]-Xo,4)
Eko = (Eko / (N * so**4) ) - 3

Xn = Xsum[1][0]/1000
Dn = Xsum[1][1]/1000 - Xn*Xn
sn = math.sqrt(Dn)
Asn = (Xsum[1][2]/1000 - 3*Xn*Xsum[1][1]/1000 + 2*Xn**3)/(sn**3)
#Ekn = (Xsum[1][3]/1000 - 4*Xn*Xsum[1][2]/1000 + 6*(Xn**2)*Xsum[1][1]/1000 - 3*Xn)/(sn**4)-3
Ekn = 0
for i in range(N):
    Ekn += math.pow(X[1][i]-Xn,4)
Ekn = (Ekn / (N * sn**4) ) - 3

Tq = Xsum[2][0]/1000
Dq = Xsum[2][1]/1000 - Tq*Tq
sq = math.sqrt(Dq)
Asq = (Xsum[2][2]/1000 - 3*Tq*Xsum[2][1]/1000 + 2*Tq**3)/(sq**3)
#Ekq = (Xsum[1][3]/1000 - 4*Tq*Xsum[1][2]/1000 + 6*(Tq**2)*Xsum[1][1]/1000 - 3*Tq)/(sq**4)-3
for i in range(N):
    Ekq += math.pow(X[2][i]-Tq,4)
Ekq = (Ekq / (N * sq**4) ) - 3

R = Xmax - Xmin
deltaX = R/10
Freq[0][0] = Xmin - deltaX
Freq[0][1] = Xmin - deltaX/2

# Расчет теретических частот
for k in range(1, 11, 1):
    Freq[k][0] = Freq[k-1][0] + deltaX
    Freq[k][1] = Freq[k-1][1] + deltaX

# Обновление массива Freq с учетом результатов моделирования каждого типа автомобилей
total_freq = Freq[:, 0].sum()  # Общая сумма частот всех автомобилей в системе
Freq[:, 2] = total_freq / Freq.shape[0]  # Общая статистика всех автомобилей в системе

# Вывод результатов
unprocessed_cars = X[1].sum()  # Суммируем количество необслуженных автомобилей в каждом эксперименте
average_unprocessed_cars = unprocessed_cars / N  # Вычисляем среднее количество необслуженных автомобилей
cel_unprocessed_cars = int(unprocessed_cars / N)  # Вычисляем целое количество необслуженных автомобилей





def calculate_mk(N, X_k):
    f_X_k = 2E-05*X_k**4 - 0.1311*X_k**3 + 394.14*X_k**2 - 526658*X_k + 3E+08
    m_k = N * f_X_k
    return m_k
# Блок обработки результатов моделирования.


# Вычисление m_k для каждого интервала
intervals = np.linspace(Xmin, Xmax, num=11)
m_k_values = np.zeros(11)
for i in range(11):
    X_k = intervals[i]
    m_k = calculate_mk(N, X_k)
    m_k_values[i] = m_k

# Вычисление критерия согласия Пирсона
chi_square = np.sum((m_k_values - Freq[:, 2])**2 / (1E+19 * Freq[:, 2]))
print("Критерий согласия Пирсона (хи-квадрат):", chi_square)

print("Среднее число необслуженных автомобилей:", average_unprocessed_cars)
print("Целое число необслуженных автомобилей:", cel_unprocessed_cars)

print("Xo = ", Xo)
print("Do = ", Do)
print("so = ", so)
print("Aso = ", Aso)
print("Eko = ", Eko)

print("Xn = ", Xn)
print("Dn = ", Dn)
print("sn = ", sn)
print("Asn = ", Asn)
print("Ekn = ", Ekn)

print("Tq = ", Tq)
print("Dq = ", Dq)
print("sq = ", sq)
print("Asq = ", Asq)
print("Ekq = ", Ekq)


# Создание нового файла Excel
# Создание нового файла Excel
workbook = Workbook()
sheet = workbook.active

# Запись значений из массива X[0] в столбец A
for i, value in enumerate(X[0], start=1):
    sheet.cell(row=i, column=1, value=value)

# Запись номеров экспериментов в столбец B
for i in range(1, N+1):
    sheet.cell(row=i, column=2, value=i)

plt.subplot(2, 2, 1)
hist, bins, _ = plt.hist(X[0], bins=40, density=False, color='blue', alpha=0.7)
plt.title('Частота автомобилей в системе')
plt.xlabel('Количество автомобилей')
plt.ylabel('Частота')

# Вычисление кривой линии тренда
x_vals = 0.5 * (bins[:-1] + bins[1:])
y_vals = hist

# Интерполяция данных для более плавной кривой
x_smooth = np.linspace(x_vals.min(), x_vals.max(), 200)
spl = make_interp_spline(x_vals, y_vals)
y_smooth = spl(x_smooth)

# Построение плавной кривой линии тренда
plt.plot(x_smooth, y_smooth, 'r-', linewidth=2)



plt.subplot(2, 2, 2)
plt.hist(X[1], bins=40, density=False, color='red', alpha=0.7)
plt.title('Частота автомобилей в очереди')
plt.xlabel('Количество автомобилей')
plt.ylabel('Частота')

plt.subplot(2, 2, 3)
plt.hist(X[2], bins=40, density=False, color='green', alpha=0.7)
plt.title('Частота автомобилей в обслуживаемых линиях')
plt.xlabel('Количество автомобилей')
plt.ylabel('Частота')

plt.subplot(2, 2, 4)
plt.hist(unprocessed_cars, bins=40, density=False, color='red', alpha=0.7)
plt.title('Частота необслуженных машин')
plt.xlabel('Количество необслуженных машин')
plt.ylabel('Частота')

plt.tight_layout()
plt.show()

#############С ЛЮБОВЬЮ ЕВГЕНИЙ###############
#############С ЛЮБОВЬЮ ЕВГЕНИЙ###############
#############С ЛЮБОВЬЮ ЕВГЕНИЙ###############
#############С ЛЮБОВЬЮ ЕВГЕНИЙ###############
#############С ЛЮБОВЬЮ ЕВГЕНИЙ###############



